import sqlite3
import uuid
from tkinter import *
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os
from tabulate import tabulate  # Import tabulate for table display
import pandas as pd


class SupplierManagement:
    def __init__(self, root):
        self.root = root
        self.root.state('zoomed')  # Set fullscreen
        self.root.title("Supplier Management")
        self.root.configure(bg="white")

        # Fonts
        self.font_title = ("Arial", 28, "bold")
        self.font_label = ("Arial", 24)
        self.font_entry = ("Arial", 22)

        # Variables
        self.supplier_name = StringVar()
        self.supplier_gst = StringVar()
        self.supplier_address = StringVar()
        self.supplier_id = StringVar(value=self.generate_supplier_id())

        # Database connection
        self.conn = sqlite3.connect('C:/Users/badla/OneDrive/Desktop/ivntmgmt/database/supplier.db')
        self.cursor = self.conn.cursor()
        self.create_table()

        # Excel file setup
        self.excel_file = "C:/Users/badla/OneDrive/Desktop/ivntmgmt/Excel/suppliers.xlsx"
        if not os.path.exists(self.excel_file):
            self.setup_excel()

        # UI setup
        self.setup_ui()

        # Load supplier data
        self.load_supplier_data()

    def create_table(self):
        """Create supplier table in the database if not exists."""
        self.cursor.execute(""" 
            CREATE TABLE IF NOT EXISTS suppliers ( 
                SUPPLIERID TEXT PRIMARY KEY, 
                NAME TEXT NOT NULL, 
                GST TEXT UNIQUE NOT NULL, 
                ADDRESS TEXT NOT NULL 
            ) 
        """)
        self.conn.commit()

    def setup_excel(self):
        """Setup Excel file with the appropriate columns."""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Suppliers"
        sheet.append(["SUPPLIERID", "NAME", "GST", "ADDRESS"])
        workbook.save(self.excel_file)

    def generate_supplier_id(self):
        """Generate a unique supplier ID."""
        return str(uuid.uuid4())[:8]

    def setup_ui(self):
        """Setup the UI components."""
        # Title
        Label(self.root, text="Supplier Management", font=self.font_title, bg="white").pack(side=TOP, pady=10)

        # Supplier Frame
        supplier_frame = LabelFrame(self.root, text="Supplier Details", font=self.font_label, bg="white", bd=2, relief=RIDGE)
        supplier_frame.place(x=20, y=80, width=self.root.winfo_screenwidth() // 3, height=self.root.winfo_screenheight() - 150)

        Label(supplier_frame, text="Name:", font=self.font_label, bg="white").grid(row=0, column=0, padx=10, pady=10, sticky=W)
        Entry(supplier_frame, textvariable=self.supplier_name, font=self.font_entry, bd=2).grid(row=0, column=1, padx=10, pady=10, sticky=W)

        Label(supplier_frame, text="GST Number:", font=self.font_label, bg="white").grid(row=1, column=0, padx=10, pady=10, sticky=W)
        Entry(supplier_frame, textvariable=self.supplier_gst, font=self.font_entry, bd=2).grid(row=1, column=1, padx=10, pady=10, sticky=W)

        Label(supplier_frame, text="Address:", font=self.font_label, bg="white").grid(row=2, column=0, padx=10, pady=10, sticky=W)
        Entry(supplier_frame, textvariable=self.supplier_address, font=self.font_entry, bd=2).grid(row=2, column=1, padx=10, pady=10, sticky=W)

        Button(supplier_frame, text="Save", font=self.font_label, bg="green", fg="white", command=self.save_supplier).grid(row=3, column=0, padx=10, pady=20, sticky=W)
        Button(supplier_frame, text="Clear", font=self.font_label, bg="blue", fg="white", command=self.clear_fields).grid(row=3, column=1, padx=10, pady=20, sticky=W)

        # Table Frame
        self.table_frame = LabelFrame(self.root, text="Supplier Table", font=self.font_label, bg="white", bd=2, relief=RIDGE)
        self.table_frame.place(x=self.root.winfo_screenwidth() // 3 + 40, y=80, width=self.root.winfo_screenwidth() * 2 // 3 - 80, height=self.root.winfo_screenheight() - 150)

        self.table = Text(self.table_frame, wrap=NONE, font=("Courier", 14, "bold"), height=25, width=100)
        self.table.pack(padx=10, pady=10)

        # Delete button
        Button(self.root, text="Delete", font=self.font_label, bg="red", fg="white", command=self.delete_supplier).pack(side=BOTTOM, pady=20)

    def save_supplier(self):
        """Save supplier details to the database and Excel."""
        name = self.supplier_name.get()
        gst = self.supplier_gst.get()
        address = self.supplier_address.get()
        supplier_id = self.supplier_id.get()

        if not name or not gst or not address:
            messagebox.showwarning("Validation Error", "Please fill in all fields.")
            return

        try:
            # Save to database
            self.cursor.execute("INSERT INTO suppliers (SUPPLIERID, NAME, GST, ADDRESS) VALUES (?, ?, ?, ?)", 
                                (supplier_id, name, gst, address))
            self.conn.commit()

            # Save to Excel
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            sheet.append([supplier_id, name, gst, address])
            workbook.save(self.excel_file)

            # Update UI with the table
            self.load_supplier_data()

            messagebox.showinfo("Success", "Supplier details saved successfully!")
            self.clear_fields()
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "GST number or Supplier ID must be unique.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def clear_fields(self):
        """Clear all input fields."""
        self.supplier_name.set("")
        self.supplier_gst.set("")
        self.supplier_address.set("")
        self.supplier_id.set(self.generate_supplier_id())

    def delete_supplier(self):
        """Delete selected supplier from database, Excel, and update table."""
        selected_item = self.table.selection_get()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a supplier to delete.")
            return

        supplier_id = selected_item.split(" | ")[0]  # Extract SupplierID from selection

        try:
            # Remove from database
            self.cursor.execute("DELETE FROM suppliers WHERE SUPPLIERID = ?", (supplier_id,))
            self.conn.commit()

            # Remove from Excel
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=False):
                if row[0].value == supplier_id:
                    sheet.delete_rows(row[0].row)
                    break
            workbook.save(self.excel_file)

            # Update table
            self.load_supplier_data()

            messagebox.showinfo("Success", "Supplier deleted successfully!")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def load_supplier_data(self):
        """Load supplier data from the database and display it using tabulate in the Text widget."""
        self.table.delete(1.0, END)  # Clear current content
        self.cursor.execute("SELECT SUPPLIERID, NAME, GST, ADDRESS FROM suppliers")
        rows = self.cursor.fetchall()

        # Create the table using tabulate
        headers = ["SUPPLIERID", "NAME", "GST", "ADDRESS"]
        table_data = tabulate(rows, headers, tablefmt="fancy_grid", numalign="center", stralign="center")

        # Insert the tabulated data into the Text widget
        self.table.insert(END, table_data)


if __name__ == "__main__":
    root = Tk()
    app = SupplierManagement(root)
    root.mainloop()
