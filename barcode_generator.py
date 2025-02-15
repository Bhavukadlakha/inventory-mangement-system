import pandas as pd
import openpyxl
from openpyxl.styles import Font
from tkinter import Tk, Button, messagebox, Toplevel, Label
from tkinter.ttk import Treeview, Scrollbar
import barcode
from barcode.writer import ImageWriter
from PIL import Image, ImageDraw, ImageFont, ImageTk
import os

class BarcodeGenerator:
    def __init__(self, root, products_file, barcode_file, barcode_photos_path):
        self.root = root
        self.products_file = products_file
        self.barcode_file = barcode_file
        self.barcode_photos_path = barcode_photos_path
        self.create_gui()

    def fetch_info(self, tree):
        try:
            # Load data from Excel, specifying columns A, B, E, and G
            products_df = pd.read_excel(self.products_file, usecols="A,B,E,G", skiprows=0)
            products_df.columns = ["Product ID", "Name", "Selling Price", "Supplier ID"]
            products_df["Barcode"] = ""  # Empty column for barcode paths
            
            # Save to barcode file
            products_df.to_excel(self.barcode_file, index=False)
            
            # Clear existing rows in the Treeview
            for row in tree.get_children():
                tree.delete(row)
            
            # Insert new rows into the Treeview
            for index, row in products_df.iterrows():
                tree.insert("", "end", values=(row["Product ID"], row["Name"], row["Selling Price"], row["Supplier ID"]))
            
            # Pre-select the first row
            if tree.get_children():
                first_item = tree.get_children()[0]
                tree.selection_set(first_item)
                tree.focus(first_item)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to fetch data: {e}")

    def generate_barcode(self, tree):
        try:
            selected_item = tree.selection()
            if not selected_item:
                raise ValueError("No product selected.")
            
            values = tree.item(selected_item, "values")
            product_id, name, selling_price, supplier_id = values[:4]
            
            # Generate barcode
            barcode_obj = barcode.get("code128", product_id, writer=ImageWriter())
            
            # Construct the correct barcode path
            barcode_path = os.path.join(self.barcode_photos_path, f"{product_id}.png")
            
            # Save the barcode image (remove extra `.png`)
            barcode_obj.save(barcode_path.replace(".png", ""))
            
            # Create a combined 2x4 inch image
            barcode_img = Image.open(barcode_path).resize((200, 100))  # Resize barcode to 2x1 inch at 200 DPI
            output_img = Image.new("RGB", (200, 200), "white")  # 2x4 inches at 200 DPI
            draw = ImageDraw.Draw(output_img)
            
            # Paste the barcode image
            output_img.paste(barcode_img, (0, 0))
            
            # Load font for text
            try:
                font_bold = ImageFont.truetype("arial.ttf", 15)
            except IOError:
                font_bold = ImageFont.load_default()
            
            # Add text below the barcode
            text_x = 10
            draw.text((text_x, 105), f"Selling Price: {selling_price}", font=font_bold, fill="black")
            draw.text((text_x, 125), f"Product ID: {product_id}", font=font_bold, fill="black")
            draw.text((text_x, 145), f"Supplier ID: {supplier_id}", font=font_bold, fill="black") 
            draw.text((text_x, 165), f"Name: {name}", font=font_bold, fill="black")
            
            # Save the final image
            output_img.save(barcode_path)
            
            # Update barcode column in Excel
            wb = openpyxl.load_workbook(self.barcode_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
                if row[0].value == product_id:
                    row[3].value = barcode_path
                    row[3].font = Font(name="3 of 9 Barcode")
            wb.save(self.barcode_file)
            
            messagebox.showinfo("Success", "Barcode generated successfully!")
            
            # Display the generated barcode
            self.display_barcode(barcode_path)
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def display_barcode(self, barcode_path):
        top = Toplevel(self.root)
        top.title("Generated Barcode")
        img = Image.open(barcode_path)
        img_tk = ImageTk.PhotoImage(img)
        label = Label(top, image=img_tk)
        label.image = img_tk
        label.pack()

    def open_file(self, tree):
        try:
            selected_item = tree.selection()
            if not selected_item:
                raise ValueError("No product selected.")
            
            values = tree.item(selected_item, "values")
            product_id = values[0]
            barcode_path = os.path.join(self.barcode_photos_path, f"{product_id}.png")  # Ensure correct '.png' extension
            
            img = Image.open(barcode_path)
            img.show()
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def create_gui(self):
        # Treeview for displaying data
        tree = Treeview(self.root, columns=("Product ID", "Name", "Selling Price", "Supplier ID", "Barcode"), show="headings", height=10)
        tree.heading("Product ID", text="Product ID")
        tree.heading("Name", text="Name")
        tree.heading("Selling Price", text="Selling Price")
        tree.heading("Supplier ID", text="Supplier ID")
        tree.heading("Barcode", text="Barcode")  

        # Scrollbar for Treeview
        scroll = Scrollbar(self.root, orient="vertical", command=tree.yview)
        tree.configure(yscroll=scroll.set)
        tree.pack(side="left", fill="both", expand=True)
        scroll.pack(side="left", fill="y")
        
        # Buttons
        fetch_button = Button(self.root, text="Fetch Info", command=lambda: self.fetch_info(tree), width=20, height=2)
        fetch_button.pack(pady=5)
        
        generate_button = Button(self.root, text="Generate Barcode", command=lambda: self.generate_barcode(tree), width=20, height=2)
        generate_button.pack(pady=5)
        
        open_file_button = Button(self.root, text="Open File", command=lambda: self.open_file(tree), width=20, height=2)
        open_file_button.pack(pady=5)

# Example usage
if __name__ == "__main__":
    root = Tk()
    products_file = r"C:\Users\badla\OneDrive\Desktop\ivntmgmt\Excel\products.xlsx"
    barcode_file = r"C:\Users\badla\OneDrive\Desktop\ivntmgmt\Excel\barcode.xlsx"
    barcode_photos_path = r"C:\Users\badla\OneDrive\Desktop\ivntmgmt\barcode photos"
    
    generator = BarcodeGenerator(root, products_file, barcode_file, barcode_photos_path)
    root.mainloop()