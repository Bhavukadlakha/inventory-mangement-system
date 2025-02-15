from tkinter import *
from tkinter import ttk, messagebox
import sqlite3
import uuid
from openpyxl import Workbook, load_workbook
import os

class ProductManagement:
    def __init__(self, root):
        self.root = root
        self.root.state('zoomed')  # Fullscreen
        self.root.title("Product Management")
        self.root.configure(bg="white")

        # Fonts
        self.font_label = ("Arial", 18)
        self.font_entry = ("Arial", 18)

        # Variables
        self.product_type = StringVar()
        self.product_quantity = IntVar(value=0)
        self.product_size = StringVar()
        self.product_cost_price = DoubleVar(value=0.0)
        self.product_selling_price = DoubleVar(value=0.0)
        self.product_gst = StringVar()
        self.supplier_gst = StringVar()
        self.supplier_name = StringVar()
        self.supplier_id = StringVar()
        self.product_id = StringVar(value=self.generate_product_id())

        # Dropdown lists
        self.types = [
            "S/Bsheet", "D/Bsheet", "S/Bblancket", "D/Bblanket", "S/Bwinterbedsheet",
            "D/Bwinterbedsheet", "S/Bmattrescover", "D/Bmatterscover", "S/Bmattres",
            "D/Bmattress", "cushion", "curtain", "cushion cover", "tsbel cover",
            "sofa cover", "blooster cover", "S/Bblanketcover", "D/Bblanketcover",
            "S/Bquilt/razai","D/Bquilt/razai", "green net", "toy", "pillow", "kid pillow", "road",
            "mat", "rug", "towel", "D/Bbedsheetset", "D/Bbookbedsheet"
        ]
        self.sizes = ["70x90", "70x100", "60x90", "90x100", "90x100", "108x108", "90x108", "93x108", "Small", "Medium", "Large", "Xtra-Large" ]
        self.gst_rates = ["5%", "8%", "12%", "18%", "28%"]

        # Files
        self.product_db = 'C:/Users/badla/OneDrive/Desktop/ivntmgmt/database/products.db'
        self.product_excel = 'C:/Users/badla/OneDrive/Desktop/ivntmgmt/Excel/products.xlsx'
        self.supplier_excel = 'C:/Users/badla/OneDrive/Desktop/ivntmgmt/Excel/supplier.xlsx'

        # Setup
        self.init_db()
        self.init_excel()
        self.setup_ui()
        self.load_products()

    def generate_product_id(self):
        """Generate a unique product ID."""
        return str(uuid.uuid4())[:8]

    def init_db(self):
        """Initialize the product database."""
        self.conn = sqlite3.connect(self.product_db)
        self.cursor = self.conn.cursor()
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS products (
                PRODUCTID TEXT PRIMARY KEY,
                TYPE TEXT NOT NULL,
                QUANTITY INTEGER NOT NULL,
                SIZE TEXT NOT NULL,
                COSTPRICE REAL NOT NULL,
                SELLINGPRICE REAL NOT NULL,
                GST TEXT NOT NULL,
                SUPPLIERID TEXT NOT NULL
            )
        """)
        self.conn.commit()

    def init_excel(self):
        """Initialize the Excel file."""
        if not os.path.exists(self.product_excel):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Products"
            sheet.append(["PRODUCTID", "TYPE", "QUANTITY", "SIZE", "SELLINGPRICE", "GST", "SUPPLIERID"])
            workbook.save(self.product_excel)

    def setup_ui(self):
        """Setup the user interface."""
        Label(self.root, text="Product Management", font=("Arial", 24, "bold"), bg="white").pack(side=TOP, pady=10)

        # Left frame: Product details
        product_frame = LabelFrame(self.root, text="Product Details", font=("Arial", 20), bg="white", bd=2, relief=RIDGE)
        product_frame.place(x=20, y=80, width=self.root.winfo_screenwidth() // 2 - 40, height=self.root.winfo_screenheight() - 150)

        # Supplier search
        Label(product_frame, text="Supplier (GST):", font=self.font_label, bg="white").grid(row=0, column=0, padx=10, pady=10, sticky=W)
        self.supplier_gst_entry = Entry(product_frame, textvariable=self.supplier_gst, font=self.font_entry, bd=2)
        self.supplier_gst_entry.grid(row=0, column=1, padx=10, pady=10, sticky=W)
        self.supplier_gst_entry.bind("<Return>", lambda event: self.search_supplier())

        Label(product_frame, text="Supplier Name:", font=self.font_label, bg="white").grid(row=1, column=0, padx=10, pady=10, sticky=W)
        self.supplier_name_entry = Entry(product_frame, textvariable=self.supplier_name, font=self.font_entry, bd=2, state='readonly')
        self.supplier_name_entry.grid(row=1, column=1, padx=10, pady=10, sticky=W)

        # Product details
        Label(product_frame, text="Type:", font=self.font_label, bg="white").grid(row=2, column=0, padx=10, pady=10, sticky=W)
        self.type_entry = ttk.Combobox(product_frame, textvariable=self.product_type, font=self.font_entry, values=self.types)
        self.type_entry.grid(row=2, column=1, padx=10, pady=10, sticky=W)

        Label(product_frame, text="Quantity:", font=self.font_label, bg="white").grid(row=3, column=0, padx=10, pady=10, sticky=W)
        self.product_quantity_entry = Entry(product_frame, textvariable=self.product_quantity, font=self.font_entry, bd=2)
        self.product_quantity_entry.grid(row=3, column=1, padx=10, pady=10, sticky=W)

        Label(product_frame, text="Size:", font=self.font_label, bg="white").grid(row=4, column=0, padx=10, pady=10, sticky=W)
        self.size_entry = ttk.Combobox(product_frame, textvariable=self.product_size, font=self.font_entry, values=self.sizes)
        self.size_entry.grid(row=4, column=1, padx=10, pady=10, sticky=W)

        Label(product_frame, text="Cost Price:", font=self.font_label, bg="white").grid(row=5, column=0, padx=10, pady=10, sticky=W)
        self.cost_price_entry = Entry(product_frame, textvariable=self.product_cost_price, font=self.font_entry, bd=2)
        self.cost_price_entry.grid(row=5, column=1, padx=10, pady=10, sticky=W)

        Label(product_frame, text="Selling Price:", font=self.font_label, bg="white").grid(row=6, column=0, padx=10, pady=10, sticky=W)
        self.selling_price_entry = Entry(product_frame, textvariable=self.product_selling_price, font=self.font_entry, bd=2)
        self.selling_price_entry.grid(row=6, column=1, padx=10, pady=10, sticky=W)

        Label(product_frame, text="GST:", font=self.font_label, bg="white").grid(row=7, column=0, padx=10, pady=10, sticky=W)
        self.gst_entry = ttk.Combobox(product_frame, textvariable=self.product_gst, font=self.font_entry, values=self.gst_rates)
        self.gst_entry.grid(row=7, column=1, padx=10, pady=10, sticky=W)

        Button(product_frame, text="Save", font=self.font_label, bg="green", fg="white", command=self.save_product).grid(row=8, column=0, padx=10, pady=20, sticky=W)
        Button(product_frame, text="Clear", font=self.font_label, bg="blue", fg="white", command=self.clear_fields).grid(row=8, column=1, padx=10, pady=20, sticky=W)

        # Right frame: Product list
        list_frame = LabelFrame(self.root, text="Product List", font=("Arial", 20), bg="white", bd=2, relief=RIDGE)
        list_frame.place(x=self.root.winfo_screenwidth() // 2, y=80, width=self.root.winfo_screenwidth() // 2 - 40, height=self.root.winfo_screenheight() - 150)

        columns = ["PRODUCTID", "TYPE", "QUANTITY", "SIZE", "SELLINGPRICE", "GST", "SUPPLIERID"]
        self.tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=20)

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150, anchor=CENTER)

        self.tree.pack(fill=BOTH, expand=True)

        Button(list_frame, text="Delete Selected", font=self.font_label, bg="red", fg="white", command=self.delete_product).pack(side=BOTTOM, pady=10)

        # Bind Enter key to move focus to the next field 
        self.supplier_gst_entry.bind("<Control-Return>", lambda event: self.search_supplier())
        self.supplier_gst_entry.bind("<Return>", lambda event: self.move_focus(self.type_entry))
        self.type_entry.bind("<Return>", lambda event: self.move_focus(self.product_quantity_entry))
        self.product_quantity_entry.bind("<Return>", lambda event: self.move_focus(self.size_entry))
        self.size_entry.bind("<Return>", lambda event: self.move_focus(self.cost_price_entry))
        self.cost_price_entry.bind("<Return>", lambda event: self.move_focus(self.selling_price_entry))
        self.selling_price_entry.bind("<Return>", lambda event: self.move_focus(self.gst_entry))
        self.gst_entry.bind("<Return>", lambda event: self.save_product())

    def move_focus(self, next_widget):
        """Move focus to the next widget."""
        next_widget.focus_set()

    def load_products(self):
        """Load all products into the TreeView."""
        for item in self.tree.get_children():
            self.tree.delete(item)

        try:
            self.cursor.execute("SELECT PRODUCTID, TYPE, QUANTITY, SIZE, SELLINGPRICE, GST, SUPPLIERID FROM products")
            rows = self.cursor.fetchall()
            for row in rows:
                self.tree.insert("", END, values=row)
        except Exception as e:
            messagebox.showerror("Error", f"Database error: {e}")

    def search_supplier(self):
        """Search supplier details from supplier.xlsx based on GST number."""
        gst = self.supplier_gst.get()
        if not gst:
            messagebox.showerror("Error", "GST number cannot be empty.")
            return

        try:
            workbook = load_workbook(self.supplier_excel)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[2] == gst:
                    self.supplier_name.set(row[1])
                    self.supplier_id.set(row[0])
                    self.type_entry.focus_set()  # Move focus to the next field
                    return
            messagebox.showerror("Error", "Supplier not found.")
        except Exception as e:
            messagebox.showerror("Error", f"Error reading supplier data: {e}")

    def save_product(self):
        """Save a new product."""
        product_id = self.product_id.get()
        product_type = self.product_type.get()
        quantity = self.product_quantity.get()
        size = self.product_size.get()
        cost_price = self.product_cost_price.get()
        selling_price = self.product_selling_price.get()
        gst = self.product_gst.get()
        supplier_id = self.supplier_id.get()

        if not all([product_type, quantity, size, cost_price, selling_price, gst, supplier_id]):
            messagebox.showerror("Error", "All fields must be filled.")
            return

        try:
            self.cursor.execute("""
                INSERT INTO products (PRODUCTID, TYPE, QUANTITY, SIZE, COSTPRICE, SELLINGPRICE, GST, SUPPLIERID)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (product_id, product_type, quantity, size, cost_price, selling_price, gst, supplier_id))
            self.conn.commit()
            self.load_products()
            self.save_to_excel(product_id, product_type, quantity, size, cost_price, selling_price, gst, supplier_id)
            self.clear_fields()
            messagebox.showinfo("Success", "Product saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Database error: {e}")

    def delete_product(self):
        """Delete the selected product."""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showerror("Error", "No product selected.")
            return

        product_id = self.tree.item(selected_item, "values")[0]

        try:
            self.cursor.execute("DELETE FROM products WHERE PRODUCTID = ?", (product_id,))
            self.conn.commit()
            self.tree.delete(selected_item)
            messagebox.showinfo("Success", "Product deleted successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Database error: {e}")

    def save_to_excel(self, product_id, product_type, quantity, size, cost_price, selling_price, gst, supplier_id):
        """Save product data to Excel."""
        try:
            workbook = load_workbook(self.product_excel)
            sheet = workbook.active
            
            # Insert a new row at the second position
            sheet.insert_rows(2)
            
            # Add the new product data to the second row, excluding cost_price
            data_to_save = [product_id, product_type, quantity, size, selling_price, gst, supplier_id]
            for col_num, value in enumerate(data_to_save, start=1):
                sheet.cell(row=2, column=col_num, value=value)
            
            workbook.save(self.product_excel)
            
            # Update the Treeview with the new product
            self.update_treeview(product_id, product_type, quantity, size, selling_price, gst, supplier_id)
        except Exception as e:
            messagebox.showerror("Error", f"Error saving to Excel: {e}")

    def update_treeview(self, product_id, product_type, quantity, size, selling_price, gst, supplier_id):
        """Update the Treeview with the new product."""
        # Insert the new product at the top of the Treeview
        self.tree.insert("", 0, values=(product_id, product_type, quantity, size, selling_price, gst, supplier_id))

    def clear_fields(self):
        """Clear all input fields."""
        self.product_id.set(self.generate_product_id())
        self.product_type.set("")
        self.product_quantity.set(0)
        self.product_size.set("")
        self.product_cost_price.set(0.0)
        self.product_selling_price.set(0.0)
        self.product_gst.set("")
        self.supplier_gst.set("")
        self.supplier_name.set("")
        self.supplier_id.set("")


if __name__ == "__main__":
    root = Tk()
    obj = ProductManagement(root)
    root.mainloop()